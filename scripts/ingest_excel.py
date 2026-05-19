#!/usr/bin/env python3
"""ingest_excel.py — Parse Excel sheet '피벗데이터(낙찰만)' and upsert into car_auctions.

Usage:
    python scripts/ingest_excel.py <path/to/file.xlsx> [--upsert]

Without --upsert the script runs in dry-run mode: parses and validates rows but
does not write to the database.

Environment variables required:
    SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY

Column mapping (0-indexed, sheet '피벗데이터(낙찰만)'):
    0  경매일자        auction_date
    1  차종            vehicle_class
    2  제작사          maker
    3  차명            car_name
    4  모델명          model_name
    5  차종(*)         full_title
    6  년식            year
    7  최초등록일      first_registered
    8  변속기          transmission
    9  엔진형식코드    engine_code
    10 배기량          engine_cc
    11 연료            fuel
    12 등급            grade
    13 색상            color
    14 주행거리        mileage_km
    15 검색구          search_category
    16 옵션            options
    17 상품구분        product_type
    18 사고품횟수      accident_count
    19 검사내용        inspection_notes
    20 시작가          start_price
    21 응찰가          bid_price
    22 낙찰금액VAT포함 final_price
    23 낙찰/유찰       status
    (col 24 is sometimes present but unused)
"""
from __future__ import annotations

import argparse
import hashlib
import os
import sys
from datetime import date, datetime, timedelta
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()

SHEET_NAME = "피벗데이터(낙찰만)"
BATCH_SIZE = 500


def _text(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _int(value) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, int):
        # Excel serial date: days since 1899-12-30 (accounting for the 1900 leap-year bug).
        try:
            return (datetime(1899, 12, 30) + timedelta(days=value)).date().strftime("%Y-%m-%d")
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def _compute_hash(auction_date, car_name, model_name, year, mileage_km, grade, color, final_price) -> str:
    """Compute a row hash that distinguishes None from empty string using a NULL sentinel (\0)."""
    parts = [auction_date, car_name, model_name, year, mileage_km, grade, color, final_price]
    norm = [("\x00" if p is None else str(p)) for p in parts]
    raw = "|".join(norm)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def parse_row(row_values: list) -> Optional[dict]:
    """Convert a list of cell values into a DB-ready dict.

    Actual column positions in '피벗데이터(낙찰만)' (verified 2026-05-19):
      0  A  개최일자          auction_date
      1  B  차군              vehicle_class (RV/중형/소형/대형 등)
      2  C  제작사            maker (예: KG모빌리티, 기아)
      3  D  차종              car_name (예: 렉스턴 스포츠, K5)
      4  E  모델명            model_name (예: 'K5 3세대(19년~현재')
      5  F  차명 (*)          full_title (제작사+차종+모델+트림 한줄)
      6  G  년식              year
      7  H  최초등록일        first_registered
      8  I  차량경력          search_category (자가용/렌터카/영업용)
      9  J  원동기형식        engine_code
      10 K  배기량            engine_cc
      11 L  연료              fuel
      12 M  평가점            grade (2글자, 예: BC)
      13 N  색상              color
      14 O  주행거리          mileage_km
      15 P  변속기            transmission
      16 Q  옵션              options
      17 R  출품구분          product_type
      18 S  재출품횟수        accident_count (재출품 횟수; 사고 횟수 X)
      19 T  사고내용          inspection_notes
      20 U  시작가            start_price
      21 V  희망가            bid_price (희망가; 응찰가 X)
      22 W  낙찰금액VAT포함   final_price ⭐
      23 X  내수/수출         (저장 안 함 — domestic_export 라벨용)

    Status is hardcoded to '낙찰' because the sheet '피벗데이터(낙찰만)' only
    contains winning bids (the name literally means 'pivot data, winners only').
    """
    if len(row_values) < 24:
        row_values = list(row_values) + [None] * (24 - len(row_values))

    car_name = _text(row_values[3])    # D 차종
    model_name = _text(row_values[4])  # E 모델명

    if not car_name or not model_name:
        return None
    if car_name == "차종" or car_name == "차명":
        return None

    auction_date = _date(row_values[0])
    grade = _text(row_values[12])      # M 평가점
    color = _text(row_values[13])      # N 색상
    mileage_km = _int(row_values[14])  # O 주행거리
    final_price = _int(row_values[22]) # W 낙찰금액VAT포함
    year = _int(row_values[6])         # G 년식

    grade_first: Optional[str] = None
    if grade and len(grade) >= 1 and grade[0].upper() in "ABCDF":
        grade_first = grade[0].upper()

    is_outlier = False
    if mileage_km is not None:
        is_outlier = mileage_km < 100 or mileage_km > 500_000

    row_hash = _compute_hash(
        auction_date, car_name, model_name, year, mileage_km, grade, color, final_price
    )

    return {
        "auction_date": auction_date,
        "vehicle_class": _text(row_values[1]),      # B 차군
        "maker": _text(row_values[2]),              # C 제작사
        "car_name": car_name,                       # D 차종
        "model_name": model_name,                   # E 모델명
        "full_title": _text(row_values[5]),         # F 차명 (*)
        "year": year,                               # G 년식
        "first_registered": _date(row_values[7]),   # H 최초등록일
        "search_category": _text(row_values[8]),    # I 차량경력 (자가용/렌터카)
        "engine_code": _text(row_values[9]),        # J 원동기형식
        "engine_cc": _int(row_values[10]),          # K 배기량
        "fuel": _text(row_values[11]),              # L 연료
        "grade": grade,                             # M 평가점
        "color": color,                             # N 색상
        "mileage_km": mileage_km,                   # O 주행거리
        "transmission": _text(row_values[15]),      # P 변속기
        "options": _text(row_values[16]),           # Q 옵션
        "product_type": _text(row_values[17]),      # R 출품구분
        "accident_count": _int(row_values[18]),    # S 재출품횟수 (semantic mismatch — kept for schema)
        "inspection_notes": _text(row_values[19]),  # T 사고내용
        "start_price": _int(row_values[20]),        # U 시작가
        "bid_price": _int(row_values[21]),          # V 희망가 (semantic mismatch — kept for schema)
        "final_price": final_price,                 # W 낙찰금액VAT포함
        "status": "낙찰",                            # Hardcoded — sheet is 낙찰만
        "grade_first": grade_first,
        "is_outlier": is_outlier,
        "row_hash": row_hash,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Excel auction data into Supabase car_auctions table.")
    parser.add_argument("path", help="Path to the .xlsx file")
    parser.add_argument("--upsert", action="store_true", help="Write to DB (default: dry-run)")
    args = parser.parse_args()

    supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

    if args.upsert and (not supabase_url or not service_key):
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set for --upsert mode.")
        sys.exit(1)

    sb = create_client(supabase_url, service_key) if args.upsert else None

    print(f"Opening workbook: {args.path}")
    wb = load_workbook(args.path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    print(f"Sheet found. Iterating rows...")

    parsed_count = 0
    skipped_count = 0
    batch: list[dict] = []
    upserted_count = 0
    dupe_count = 0
    warn_count = 0
    batches_ok = 0
    batches_fail = 0
    batch_index = 0

    def _flush_batch(b: list[dict]) -> None:
        nonlocal upserted_count, dupe_count, batches_ok, batches_fail, batch_index
        batch_index += 1
        if not sb:
            upserted_count += len(b)
            batches_ok += 1
            return
        try:
            resp = sb.table("car_auctions").upsert(b, on_conflict="row_hash", ignore_duplicates=True).execute()
            inserted = len(resp.data) if resp.data else 0
            dupes = len(b) - inserted
            upserted_count += inserted
            dupe_count += dupes
            batches_ok += 1
        except Exception as exc:
            first_hash = b[0].get("row_hash", "?") if b else "?"
            print(f"  [WARN] Batch {batch_index} FAILED (first row_hash={first_hash}): {exc}")
            batches_fail += 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = parse_row(list(row))
        if record is None:
            skipped_count += 1
            continue

        if record["auction_date"] is None:
            warn_count += 1
            skipped_count += 1
            continue

        parsed_count += 1
        batch.append(record)

        if len(batch) >= BATCH_SIZE:
            _flush_batch(batch)
            batch = []

        if parsed_count % 5000 == 0:
            print(f"  Parsed {parsed_count:,} rows so far...")

    # Flush remainder
    if batch:
        _flush_batch(batch)

    wb.close()

    total_batches = batches_ok + batches_fail

    print()
    print(f"=== Ingest complete ({'DRY-RUN' if not args.upsert else 'LIVE'}) ===")
    print(f"  Rows parsed:    {parsed_count:,}")
    print(f"  Rows skipped:   {skipped_count:,}  (empty/header)")
    print(f"  Warnings:       {warn_count:,}  (missing auction_date)")
    print(f"  Batches:        {batches_ok}/{total_batches} succeeded, {batches_fail} failed")
    if args.upsert:
        print(f"  Upserted (new): {upserted_count:,}")
        print(f"  Duplicates:     {dupe_count:,}  (ON CONFLICT DO NOTHING)")
        if dupe_count > parsed_count * 0.9:
            print("  [ALERT] More than 90% duplicates — check if this file was already ingested.")
        if batches_fail:
            print(f"  [ALERT] {batches_fail} batch(es) failed — some rows were NOT upserted.")


if __name__ == "__main__":
    main()
