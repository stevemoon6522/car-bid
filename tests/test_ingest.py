"""tests/test_ingest.py — Smoke tests for ingest_excel.py parsing logic.

Creates a tiny synthetic Excel in-memory (5 rows) and exercises parse_row
without touching the database or the 42MB production file.
"""
import io
import sys
import os

import pytest

# Make sure the scripts directory is importable.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))

import openpyxl
from ingest_excel import parse_row, _compute_hash, _text, _int, _date


# ---------------------------------------------------------------------------
# Helper: build a minimal row (24 values matching the column mapping)
# ---------------------------------------------------------------------------

def _make_values(
    auction_date="2025-05-21",
    vehicle_class="SUV",
    maker="기아",
    car_name="K5",
    model_name="K5 3세대(19년~현재",
    full_title="K5 3세대 2.0 GDI",
    year=2021,
    first_registered="2021-06-18",
    transmission="자동",
    engine_code="G4NA",
    engine_cc=1999,
    fuel="가솔린",
    grade="BB",
    color="검정",
    mileage_km=80000,
    search_category="국산",
    options="네비,스마트키,선루프",
    product_type="정상매물",
    accident_count=0,
    inspection_notes="이상없음",
    start_price=1500,
    bid_price=1580,
    final_price=1580,
    status="낙찰",
):
    return [
        auction_date, vehicle_class, maker, car_name, model_name, full_title,
        year, first_registered, transmission, engine_code, engine_cc, fuel,
        grade, color, mileage_km, search_category, options, product_type,
        accident_count, inspection_notes, start_price, bid_price, final_price, status,
    ]


# ---------------------------------------------------------------------------
# parse_row tests
# ---------------------------------------------------------------------------

class TestParseRow:
    def test_valid_row_returns_dict(self):
        record = parse_row(_make_values())
        assert record is not None
        assert record["car_name"] == "K5"
        assert record["model_name"] == "K5 3세대(19년~현재"
        assert record["status"] == "낙찰"

    def test_grade_first_extracted(self):
        record = parse_row(_make_values(grade="BC"))
        assert record["grade_first"] == "B"

    def test_grade_first_uppercase(self):
        record = parse_row(_make_values(grade="ab"))
        assert record["grade_first"] == "A"

    def test_grade_first_f_is_valid(self):
        # F appears in grade but the second-position meaning; first position F → valid char
        record = parse_row(_make_values(grade="FB"))
        assert record["grade_first"] == "F"

    def test_outlier_flag_high_mileage(self):
        record = parse_row(_make_values(mileage_km=600_000))
        assert record["is_outlier"] is True

    def test_outlier_flag_low_mileage(self):
        record = parse_row(_make_values(mileage_km=50))
        assert record["is_outlier"] is True

    def test_normal_mileage_not_outlier(self):
        record = parse_row(_make_values(mileage_km=80_000))
        assert record["is_outlier"] is False

    def test_missing_car_name_returns_none(self):
        values = _make_values(car_name=None)
        assert parse_row(values) is None

    def test_missing_model_name_returns_none(self):
        values = _make_values(model_name=None)
        assert parse_row(values) is None

    def test_status_hardcoded_to_won(self):
        # The sheet '피벗데이터(낙찰만)' contains only winning bids; parse_row
        # therefore ignores the input status column (col X = 내수/수출) and
        # always emits status='낙찰'. Verify the behavior — passing
        # status=None / "anything" should NOT cause the row to be skipped.
        for s in (None, "수출", "내수", "유찰", ""):
            values = _make_values(status=s)
            record = parse_row(values)
            assert record is not None, f"Row dropped unexpectedly when status={s!r}"
            assert record["status"] == "낙찰"

    def test_header_row_skipped(self):
        # A header row where car_name cell literally says "차종" (D column header)
        for hdr in ("차종", "차명"):
            values = _make_values(car_name=hdr)
            assert parse_row(values) is None

    def test_row_hash_is_sha256_hex(self):
        record = parse_row(_make_values())
        assert record["row_hash"] is not None
        assert len(record["row_hash"]) == 64

    def test_row_hash_deterministic(self):
        v1 = _make_values()
        v2 = _make_values()
        assert parse_row(v1)["row_hash"] == parse_row(v2)["row_hash"]

    def test_row_hash_differs_on_price_change(self):
        v1 = _make_values(final_price=1580)
        v2 = _make_values(final_price=1600)
        assert parse_row(v1)["row_hash"] != parse_row(v2)["row_hash"]

    def test_short_row_padded_gracefully(self):
        # Row with only 10 values should not raise IndexError. car_name (idx 3)
        # and model_name (idx 4) are present, so parse_row returns a record;
        # later fields (mileage, prices, etc.) are simply None.
        short = _make_values()[:10]
        record = parse_row(short)
        assert record is not None
        assert record["car_name"] == "K5"
        assert record["model_name"] == "K5 3세대(19년~현재"
        assert record["status"] == "낙찰"
        assert record["final_price"] is None
        assert record["mileage_km"] is None

    def test_integer_fields_parsed(self):
        record = parse_row(_make_values(year=2020, mileage_km=50_000, final_price=1200))
        assert record["year"] == 2020
        assert record["mileage_km"] == 50_000
        assert record["final_price"] == 1200

    def test_none_final_price_handled(self):
        record = parse_row(_make_values(final_price=None, status="유찰"))
        assert record is not None
        assert record["final_price"] is None


# ---------------------------------------------------------------------------
# Synthetic Excel end-to-end test
# ---------------------------------------------------------------------------

class TestSyntheticExcel:
    """Build a tiny in-memory workbook and verify parse_row handles all 5 rows."""

    ROWS = [
        _make_values(),                                          # normal 낙찰
        _make_values(status="유찰"),                             # 유찰 row
        _make_values(mileage_km=600_000),                        # outlier high
        _make_values(mileage_km=50),                             # outlier low
        _make_values(car_name="차명"),                            # header-like
    ]

    def _build_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "이번주거래(누적)"
        # Header row
        ws.append(["경매일자", "차종", "제작사", "차명", "모델명", "차종(*)",
                   "년식", "최초등록일", "변속기", "엔진형식코드", "배기량", "연료",
                   "등급", "색상", "주행거리", "검색구", "옵션", "상품구분",
                   "사고품횟수", "검사내용", "시작가", "응찰가", "낙찰금액VAT포함", "낙찰/유찰"])
        for row in self.ROWS:
            ws.append(row)
        return wb

    def test_parses_five_rows(self):
        wb = self._build_workbook()
        ws = wb["이번주거래(누적)"]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            r = parse_row(list(row))
            results.append(r)
        # 5 data rows total
        assert len(results) == 5

    def test_header_like_row_yields_none(self):
        wb = self._build_workbook()
        ws = wb["이번주거래(누적)"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # Last row has car_name="차명"
        assert parse_row(list(rows[-1])) is None

    def test_valid_rows_count(self):
        wb = self._build_workbook()
        ws = wb["이번주거래(누적)"]
        parsed = [
            parse_row(list(row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        valid = [r for r in parsed if r is not None]
        # 5 rows: 4 valid (including 유찰 and outliers), 1 header-like → None
        assert len(valid) == 4

    def test_outlier_flags_set(self):
        wb = self._build_workbook()
        ws = wb["이번주거래(누적)"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        high_mileage = parse_row(list(rows[2]))
        low_mileage = parse_row(list(rows[3]))
        assert high_mileage["is_outlier"] is True
        assert low_mileage["is_outlier"] is True


# ---------------------------------------------------------------------------
# M8: _date helper — Excel serial int conversion
# ---------------------------------------------------------------------------

class TestDateHelper:
    def test_none_returns_none(self):
        assert _date(None) is None

    def test_empty_string_returns_none(self):
        assert _date("") is None

    def test_datetime_object(self):
        from datetime import datetime
        assert _date(datetime(2025, 5, 21)) == "2025-05-21"

    def test_date_object(self):
        from datetime import date
        assert _date(date(2025, 5, 21)) == "2025-05-21"

    def test_string_dash(self):
        assert _date("2025-05-21") == "2025-05-21"

    def test_string_slash(self):
        assert _date("2025/05/21") == "2025-05-21"

    def test_string_dot(self):
        assert _date("2025.05.21") == "2025-05-21"

    def test_excel_serial_int(self):
        # Excel serial 45798 = 2025-05-21 (days since 1899-12-30).
        from datetime import datetime, timedelta
        expected = (datetime(1899, 12, 30) + timedelta(days=45798)).date().strftime("%Y-%m-%d")
        assert _date(45798) == expected

    def test_invalid_string_returns_none(self):
        assert _date("not-a-date") is None


# ---------------------------------------------------------------------------
# M9: _compute_hash — None vs empty string distinctness
# ---------------------------------------------------------------------------

class TestComputeHash:
    def test_none_differs_from_empty_string(self):
        h1 = _compute_hash(None, "K5", "K5 3세대", 2021, 80000, "BB", "검정", 1580)
        h2 = _compute_hash("", "K5", "K5 3세대", 2021, 80000, "BB", "검정", 1580)
        assert h1 != h2

    def test_all_none_fields_deterministic(self):
        h1 = _compute_hash(None, None, None, None, None, None, None, None)
        h2 = _compute_hash(None, None, None, None, None, None, None, None)
        assert h1 == h2

    def test_different_fields_differ(self):
        h1 = _compute_hash("2025-05-01", "K5", "K5 3세대", 2021, 80000, "BB", "검정", 1580)
        h2 = _compute_hash("2025-05-01", "K5", "K5 3세대", 2021, 80000, "BB", "검정", 1600)
        assert h1 != h2


# ---------------------------------------------------------------------------
# M7: _parse_korean_int — tested via app.py helper
# ---------------------------------------------------------------------------

class TestParseKoreanInt:
    def setup_method(self):
        # Import the helper from app.py (it's in the project root).
        import importlib, sys, os
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
        import app as _app
        self._fn = _app._parse_korean_int

    def test_plain_int(self):
        assert self._fn("80000", "주행거리") == 80000

    def test_comma_separated(self):
        assert self._fn("80,000", "주행거리") == 80000

    def test_man_suffix(self):
        assert self._fn("8만", "주행거리") == 80000

    def test_man_suffix_with_comma(self):
        assert self._fn("8만", "주행거리") == 80000

    def test_year_plain(self):
        assert self._fn("2021", "년식") == 2021

    def test_invalid_raises_400(self):
        from fastapi import HTTPException
        with pytest.raises(HTTPException) as exc_info:
            self._fn("abc", "주행거리")
        assert exc_info.value.status_code == 400

    def test_empty_raises_400(self):
        from fastapi import HTTPException
        with pytest.raises(HTTPException):
            self._fn("", "주행거리")
